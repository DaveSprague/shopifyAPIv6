#!/usr/bin/env python3
"""
Beautiful Excel Mismatch Analysis
Creates formatted Excel workbook with color coding for payment reconciliation mismatches
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
import subprocess

# Configuration settings
DATE_RANGE_DAYS = 120  # Number of days to show (0 = show all)
START_DATE = "2025-01-01"  # Default starting date (None = use earliest mismatch date, format: "2024-01-01")
REVERSE_DATE_ORDER = False  # True = newest first, False = oldest first
MAX_COLUMNS = 20  # Maximum number of date columns to show (0 = no limit)

def filter_and_sort_dates(mismatch_dates, reverse_order=REVERSE_DATE_ORDER, 
                         start_date=START_DATE, max_days=DATE_RANGE_DAYS, 
                         max_columns=MAX_COLUMNS):
    """Filter and sort mismatch dates based on configuration"""
    
    if not mismatch_dates:
        return []
    
    # Convert to datetime objects for proper sorting
    date_objects = []
    for date_str in mismatch_dates:
        try:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d")
            date_objects.append((date_obj, date_str))
        except ValueError:
            print(f"⚠️  Skipping invalid date format: {date_str}")
    
    # Sort by date
    date_objects.sort(key=lambda x: x[0], reverse=reverse_order)
    
    # Apply start date filter if specified
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            if reverse_order:
                # For reverse order, start_date is the latest date to show
                date_objects = [d for d in date_objects if d[0] <= start_dt]
            else:
                # For forward order, start_date is the earliest date to show
                date_objects = [d for d in date_objects if d[0] >= start_dt]
        except ValueError:
            print(f"⚠️  Invalid start date format: {start_date}, using all dates")
    
    # Apply max days filter
    if max_days > 0 and date_objects:
        if reverse_order:
            # For reverse order, count backwards from the start
            reference_date = date_objects[0][0]  # Latest date
            cutoff_date = reference_date - timedelta(days=max_days)
            date_objects = [d for d in date_objects if d[0] >= cutoff_date]
        else:
            # For forward order, count forwards from the start
            reference_date = date_objects[0][0]  # Earliest date
            cutoff_date = reference_date + timedelta(days=max_days)
            date_objects = [d for d in date_objects if d[0] <= cutoff_date]
    
    # Apply max columns limit
    if max_columns > 0:
        date_objects = date_objects[:max_columns]
    
    # Extract date strings
    filtered_dates = [date_str for _, date_str in date_objects]
    
    print(f"📅 Date filtering applied:")
    print(f"   Order: {'Newest → Oldest' if reverse_order else 'Oldest → Newest'}")
    print(f"   Start date: {start_date if start_date else 'Auto (first mismatch)'}")
    print(f"   Max days: {max_days if max_days > 0 else 'All'}")
    print(f"   Max columns: {max_columns if max_columns > 0 else 'All'}")
    print(f"   Filtered: {len(mismatch_dates)} → {len(filtered_dates)} dates")
    
    if filtered_dates:
        print(f"   Range: {filtered_dates[-1] if reverse_order else filtered_dates[0]} to {filtered_dates[0] if reverse_order else filtered_dates[-1]}")
    
    return filtered_dates

def show_date_configuration():
    """Display current date configuration settings"""
    print("\n📅 CURRENT DATE CONFIGURATION")
    print("=" * 40)
    print(f"Date Order: {'Newest → Oldest' if REVERSE_DATE_ORDER else 'Oldest → Newest'}")
    print(f"Start Date: {START_DATE if START_DATE else 'Auto (first mismatch)'}")
    print(f"Max Days: {DATE_RANGE_DAYS if DATE_RANGE_DAYS > 0 else 'All'}")
    print(f"Max Columns: {MAX_COLUMNS if MAX_COLUMNS > 0 else 'All'}")
    print()

def configure_date_settings():
    """Interactive configuration of date settings"""
    global REVERSE_DATE_ORDER, START_DATE, DATE_RANGE_DAYS, MAX_COLUMNS
    
    print("\n⚙️  DATE CONFIGURATION")
    print("=" * 30)
    
    # Date order
    order_choice = input(f"Date order (1=Oldest→Newest, 2=Newest→Oldest) [Current: {'2' if REVERSE_DATE_ORDER else '1'}]: ").strip()
    if order_choice == "1":
        REVERSE_DATE_ORDER = False
    elif order_choice == "2":
        REVERSE_DATE_ORDER = True
    
    # Start date
    start_choice = input(f"Start date (YYYY-MM-DD or Enter for auto) [Current: {START_DATE or 'Auto'}]: ").strip()
    if start_choice:
        try:
            datetime.strptime(start_choice, "%Y-%m-%d")
            START_DATE = start_choice
        except ValueError:
            print("❌ Invalid date format, keeping current setting")
    elif start_choice == "":
        START_DATE = "2025-04-01"  # Default to recent date instead of None
    
    # Max days
    days_choice = input(f"Max days to show (0=all) [Current: {DATE_RANGE_DAYS}]: ").strip()
    if days_choice:
        try:
            DATE_RANGE_DAYS = max(0, int(days_choice))
        except ValueError:
            print("❌ Invalid number, keeping current setting")
    
    # Max columns
    cols_choice = input(f"Max columns to show (0=all) [Current: {MAX_COLUMNS}]: ").strip()
    if cols_choice:
        try:
            MAX_COLUMNS = max(0, int(cols_choice))
        except ValueError:
            print("❌ Invalid number, keeping current setting")
    
    print("\n✅ Configuration updated!")
    show_date_configuration()

def create_beautiful_excel_table():
    """Create beautiful Excel file with color coding and formatting"""
    
    try:
        # Load transposed reconciliation data (shop timezone for consistent comparison)
        df = pd.read_csv('transposed_reconciliation_shop_timezone.csv', index_col=0)
        print("📊 BEAUTIFUL EXCEL MISMATCH ANALYSIS (SHOP TIMEZONE)")
        print("=" * 60)
        print("🕐 Using shop timezone data for consistency with Shopify reports")
        
        # Debug shopify_payments issue
        print(f"\n🔍 DEBUGGING SHOPIFY_PAYMENTS DATA:")
        if 'shopify_payments' in df.index:
            shopify_row = df.loc['shopify_payments']
            print(f"   Total columns: {len(shopify_row)}")
            print(f"   Non-zero values: {(shopify_row != 0).sum()}")
            print(f"   Sample values: {shopify_row.head(10).to_dict()}")
            
            # Check if all values are zero
            if (shopify_row == 0).all():
                print("   ⚠️  ALL shopify_payments values are zero!")
                print("   💡 This suggests an issue with payment gateway data extraction")
            else:
                non_zero_dates = shopify_row[shopify_row != 0].index.tolist()
                print(f"   ✅ Found non-zero shopify_payments on {len(non_zero_dates)} dates")
                if non_zero_dates:
                    print(f"   📅 Sample non-zero dates: {non_zero_dates[:5]}")
        else:
            print("   ❌ 'shopify_payments' row not found in data!")
        
        # Also check for payment gateway data
        payment_gateways = ['shopify_payments', 'gift_card', 'cash', 'manual']
        print(f"\n🔍 PAYMENT GATEWAY SUMMARY:")
        for gateway in payment_gateways:
            if gateway in df.index:
                gateway_row = df.loc[gateway]
                # Convert to numeric values for proper calculation
                gateway_numeric = pd.to_numeric(gateway_row, errors='coerce').fillna(0)
                non_zero_count = (gateway_numeric != 0).sum()
                total_value = gateway_numeric.sum()
                print(f"   {gateway}: {non_zero_count} non-zero days, total: ${total_value:,.2f}")
            else:
                print(f"   {gateway}: NOT FOUND")
        print()
        
        # Get mismatch dates
        mismatch_row = df.loc['mismatch'] if 'mismatch' in df.index else None
        if mismatch_row is None:
            print("❌ No 'mismatch' row found in data")
            return
            
        mismatch_dates = [col for col in df.columns if str(mismatch_row[col]).lower() == 'true']
        
        if not mismatch_dates:
            print("✅ No mismatches found! All dates reconcile properly.")
            return
        
        # Apply date filtering and sorting
        print(f"📊 Found {len(mismatch_dates)} mismatched dates")
        filtered_dates = filter_and_sort_dates(mismatch_dates)
        
        if not filtered_dates:
            print("❌ No dates remain after filtering!")
            return
        
        print("📝 Creating beautifully formatted Excel file...")
        
        # Use filtered dates
        mismatch_dates = filtered_dates
        
        # Create focused DataFrame with only mismatch dates
        focus_rows = [
            # Sales components
            'order_count',
            'gross_sales',
            'subtotal',
            'discounts',
            'tax',
            'shipping', 
            'tips',
            'net_payment',
            'refunds',
            # Payment gateways
            'shopify_payments',
            'gift_card',
            'cash', 
            'manual',
            # Payment adjustments from payout data
            'payout_type_adjustment',
            'payout_type_chargeback',
            'payout_type_refund',
            # Payout information
            'payout_amount',
            'reconciliation_difference',
            'pending_payout_amount',
            'pending_payout_count'
        ]
        
        # Filter to only rows we care about and only mismatch dates
        available_rows = [row for row in focus_rows if row in df.index]
        mismatch_df = df.loc[available_rows, mismatch_dates]
        
        # Keep data transposed so dates are columns and metrics are rows
        # Convert to numeric values and handle NaN for each date column
        for date_col in mismatch_dates:
            if date_col in mismatch_df.columns:
                for idx in mismatch_df.index:
                    try:
                        mismatch_df.loc[idx, date_col] = pd.to_numeric(mismatch_df.loc[idx, date_col], errors='coerce')
                        if pd.isna(mismatch_df.loc[idx, date_col]):
                            mismatch_df.loc[idx, date_col] = 0
                    except:
                        mismatch_df.loc[idx, date_col] = 0
        
        # Add calculated rows with proper sales balance calculations
        # Sales calculation: gross_sales - discounts + tax + shipping + tips should equal net_payment + refunds
        sales_total_row = (mismatch_df.loc['gross_sales'] - 
                          mismatch_df.loc['discounts'] + 
                          mismatch_df.loc['tax'] + 
                          mismatch_df.loc['shipping'] + 
                          mismatch_df.loc['tips'])
        
        expected_from_sales_row = mismatch_df.loc['net_payment'] + mismatch_df.loc['refunds']
        sales_balance_check_row = sales_total_row - expected_from_sales_row
        
        # Payment gateway totals including adjustments
        gateway_payments_row = (
            mismatch_df.loc['shopify_payments'] + 
            mismatch_df.loc['gift_card'] + 
            mismatch_df.loc['cash'] + 
            mismatch_df.loc['manual']
        )
        
        # Payment adjustments (note: payout_type_refund is already negative)
        payment_adjustments_row = (
            mismatch_df.loc['payout_type_adjustment'] + 
            mismatch_df.loc['payout_type_chargeback'] + 
            mismatch_df.loc['payout_type_refund']  # Already negative in the data
        )
        
        # For proper reconciliation: compare total shopify receipts with expected payout before fees
        # Total Shopify receipts = shopify_payments + payout_refunds (since refunds were originally part of payments)
        payout_charges_row = mismatch_df.loc['payout_amount']  # Charges from payout data (before fees)
        payout_refunds_row = abs(mismatch_df.loc['payout_type_refund'])  # Refunds (make positive for clarity)
        expected_payout_before_fees_row = payout_charges_row  # This is the expected payout before fees
        
        # Total Shopify payment receipts includes the original payments plus refunds
        total_shopify_receipts_row = mismatch_df.loc['shopify_payments'] + payout_refunds_row
        
        total_gateway_receipts_row = gateway_payments_row + payment_adjustments_row
        payment_difference_row = total_shopify_receipts_row - expected_payout_before_fees_row
        
        # Add these as new rows to the DataFrame
        mismatch_df.loc['sales_total'] = sales_total_row
        mismatch_df.loc['expected_from_sales'] = expected_from_sales_row  
        mismatch_df.loc['sales_balance_check'] = sales_balance_check_row
        mismatch_df.loc['empty_row'] = 0  # Spacer row
        mismatch_df.loc['gateway_payments'] = gateway_payments_row
        mismatch_df.loc['payment_adjustments'] = payment_adjustments_row
        mismatch_df.loc['total_gateway_receipts'] = total_gateway_receipts_row
        mismatch_df.loc['payout_charges'] = payout_charges_row
        mismatch_df.loc['payout_refunds_amount'] = payout_refunds_row
        mismatch_df.loc['total_shopify_receipts'] = total_shopify_receipts_row
        mismatch_df.loc['expected_payout_before_fees'] = expected_payout_before_fees_row
        mismatch_df.loc['payment_difference'] = payment_difference_row
        
        # Reorder rows for better Excel layout with logical grouping
        excel_rows = [
            # Sales Analysis Section
            'order_count',
            'gross_sales',
            'subtotal',
            'discounts', 
            'tax',
            'shipping',
            'tips',
            'sales_total',
            'net_payment',
            'refunds',
            'expected_from_sales',
            'sales_balance_check',
            'empty_row',  # Visual separator
            # Payment Gateway Section
            'shopify_payments',
            'gift_card', 
            'cash',
            'manual',
            'gateway_payments',
            'payout_type_adjustment',
            'payout_type_chargeback',
            'payout_type_refund', 
            'payment_adjustments',
            'total_gateway_receipts',
            'empty_row',  # Visual separator for payout reconciliation
            # Payout Reconciliation Section
            'payout_amount',
            'payout_charges',
            'payout_refunds_amount',
            'total_shopify_receipts',
            'expected_payout_before_fees',
            'payment_difference',
            'reconciliation_difference',
            'pending_payout_amount',
            'pending_payout_count'
        ]
        
        # Filter to existing rows and sort dates chronologically
        available_rows = [row for row in excel_rows if row in mismatch_df.index]
        mismatch_dates_sorted = mismatch_dates  # Already sorted by filter function
        excel_df = mismatch_df.loc[available_rows, mismatch_dates_sorted].copy()
        
        # Create Excel file with formatting
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f'Payment_Reconciliation_Mismatches_ShopTimezone_{timestamp}.xlsx'
        
        # Try to use the standard filename first, fall back to timestamped version
        try:
            create_formatted_excel(excel_df, 'Payment_Reconciliation_Mismatches_ShopTimezone.xlsx')
            excel_filename = 'Payment_Reconciliation_Mismatches_ShopTimezone.xlsx'
        except PermissionError:
            print("⚠️  Previous Excel file is open, creating new timestamped version...")
            create_formatted_excel(excel_df, excel_filename)
        
        print(f"✅ Created: {excel_filename}")
        print("🚀 Opening in Excel...")
        
        # Open in Excel
        open_excel_file(excel_filename)
        
    except FileNotFoundError:
        print("❌ transposed_reconciliation_shop_timezone.csv not found")
        print("💡 Run reconcile_payouts.py first to generate the shop timezone data")
        print("🕐 This analysis uses shop timezone data for consistency with Shopify reports")
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()

def create_formatted_excel(df, filename):
    """Create beautifully formatted Excel file with color coding"""
    
    try:
        # Try to use openpyxl for better formatting
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Write main data
            df.to_excel(writer, sheet_name='Mismatch Analysis', startrow=2)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Mismatch Analysis']
            
            # Apply formatting
            format_excel_worksheet(workbook, worksheet, df)
            
    except ImportError:
        # Fallback to xlsxwriter if openpyxl not available
        try:
            with pd.ExcelWriter(filename, engine='xlsxwriter') as writer:
                df.to_excel(writer, sheet_name='Mismatch Analysis', startrow=2)
                
                workbook = writer.book
                worksheet = writer.sheets['Mismatch Analysis']
                
                format_excel_worksheet_xlsxwriter(workbook, worksheet, df)
                
        except ImportError:
            # Basic Excel export without formatting
            df.to_excel(filename, sheet_name='Mismatch Analysis')
            print("⚠️  Basic Excel file created (install openpyxl or xlsxwriter for formatting)")

def format_excel_worksheet(workbook, worksheet, df):
    """Format Excel worksheet using openpyxl with dates as columns"""
    
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Define colors
    header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    high_diff_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    medium_diff_fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
    low_diff_fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
    
    # Define fonts
    header_font = Font(color="FFFFFF", bold=True, size=11)
    bold_font = Font(bold=True)
    
    # Define alignment
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    # Define border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title
    title_cell = worksheet['A1']
    title_cell.value = "🔍 Payment Reconciliation Mismatches Analysis (Shop Timezone)"
    title_cell.font = Font(size=16, bold=True, color="2E4057")
    
    # Add timezone info in row 2
    timezone_cell = worksheet['A2']
    timezone_cell.value = "🕐 Data aligned to shop timezone for consistency with Shopify reports"
    timezone_cell.font = Font(size=10, italic=True, color="666666")
    
    # Merge title across all columns
    last_col = get_column_letter(len(df.columns) + 1)
    worksheet.merge_cells(f'A1:{last_col}1')
    worksheet.merge_cells(f'A2:{last_col}2')
    
    # Format date headers (row 3, columns B onwards)
    header_row = 3
    for col_num, date_col in enumerate(df.columns, 2):  # Start at column B
        cell = worksheet.cell(row=header_row, column=col_num)
        cell.value = date_col
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # Format metric row headers (column A, starting row 4)
    metric_labels = {
        # Sales Analysis
        'order_count': 'Order Count',
        'gross_sales': 'Gross Sales',
        'subtotal': 'Subtotal',
        'discounts': 'Discounts',
        'tax': 'Tax',
        'shipping': 'Shipping',
        'tips': 'Tips',
        'sales_total': '📊 Sales Total (Gross - Disc + Tax + Ship + Tips)',
        'net_payment': 'Net Payment Expected',
        'refunds': 'Refunds',
        'expected_from_sales': '🔍 Expected from Sales (Net + Refunds)',
        'sales_balance_check': '⚠️ Sales Balance Check',
        'empty_row': '',
        # Payment Gateways
        'shopify_payments': 'Shopify Payments',
        'gift_card': 'Gift Cards',
        'cash': 'Cash',
        'manual': 'Manual',
        'gateway_payments': '💳 Gateway Payments Subtotal',
        'payout_type_adjustment': 'Payout Adjustments',
        'payout_type_chargeback': 'Chargebacks',
        'payout_type_refund': 'Payout Refunds',
        'payment_adjustments': '🔧 Payment Adjustments Subtotal',
        'total_gateway_receipts': '📈 Total Gateway Receipts',
        # Payout Reconciliation
        'payout_amount': 'Payout Charges',
        'payout_charges': 'Payout Charges (same as above)',
        'payout_refunds_amount': 'Payout Refunds (positive)',
        'expected_payout_net': '🧮 Expected Payout Net (Charges - Refunds)',
        'payment_difference': '❗ Payment Difference (Shopify - Expected)',
        'reconciliation_difference': 'Reconciliation Difference',
        'pending_payout_amount': 'Pending Amount',
        'pending_payout_count': 'Pending Count'
    }
    
    # Format metric name column header
    metric_header = worksheet.cell(row=header_row, column=1)
    metric_header.value = 'Metric'
    metric_header.fill = header_fill
    metric_header.font = header_font
    metric_header.alignment = right_align  # Changed from center_align to right_align
    metric_header.border = thin_border
    
    # Format data rows
    for row_num, (metric, row_data) in enumerate(df.iterrows(), header_row + 1):
        # Handle empty spacer row
        if metric == 'empty_row':
            for col_num in range(1, len(df.columns) + 2):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = ''
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            continue
            
        # Metric name column
        metric_cell = worksheet.cell(row=row_num, column=1)
        metric_cell.value = metric_labels.get(metric, metric.replace('_', ' ').title())
        metric_cell.alignment = right_align  # Changed from center_align to right_align
        metric_cell.border = thin_border
        
        # Determine special formatting for different row types
        is_summary_metric = metric in ['sales_total', 'expected_from_sales', 'gateway_payments', 
                                     'payment_adjustments', 'total_gateway_receipts', 'expected_payout_net']
        is_check_metric = metric in ['sales_balance_check', 'payment_difference']
        is_section_header = metric in ['order_count']  # First row of each section
        
        # Set metric cell formatting
        if is_summary_metric:
            metric_cell.font = Font(bold=True, color="1F4E79")
            metric_cell.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
        elif is_check_metric:
            metric_cell.font = Font(bold=True, color="C5504B")  
            metric_cell.fill = PatternFill(start_color="FFEEEE", end_color="FFEEEE", fill_type="solid")
        else:
            metric_cell.font = bold_font
            metric_cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        # Format data columns
        for col_num, date_col in enumerate(df.columns, 2):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell_value = row_data[date_col]
            cell.value = cell_value
            cell.border = thin_border
            
            # Determine cell color based on different criteria
            if metric == 'payment_difference':
                payment_diff = abs(float(cell_value)) if cell_value != 0 else 0
                if payment_diff > 100:
                    cell.fill = high_diff_fill
                elif payment_diff > 50:
                    cell.fill = medium_diff_fill
                else:
                    cell.fill = low_diff_fill
            elif metric == 'sales_balance_check':
                # Highlight sales balance issues
                balance_diff = abs(float(cell_value)) if cell_value != 0 else 0
                if balance_diff > 1:  # More than $1 difference
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            elif is_summary_metric:
                cell.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
            elif is_check_metric:
                cell.fill = PatternFill(start_color="FFEEEE", end_color="FFEEEE", fill_type="solid")
            
            # Number formatting and alignment
            if metric in ['order_count', 'pending_payout_count']:
                cell.alignment = center_align
                cell.number_format = '0'
            elif metric in ['gross_sales', 'subtotal', 'discounts', 'tax', 'shipping', 'tips', 'sales_total',
                          'net_payment', 'refunds', 'expected_from_sales', 'sales_balance_check',
                          'shopify_payments', 'gift_card', 'cash', 'manual', 'gateway_payments',
                          'payout_type_adjustment', 'payout_type_chargeback', 'payout_type_refund', 'payment_adjustments',
                          'total_gateway_receipts', 'payment_difference',
                          'payout_amount', 'payout_charges', 'payout_refunds_amount', 'expected_payout_net',
                          'total_shopify_receipts', 'expected_payout_before_fees',  # Added new metrics
                          'reconciliation_difference', 'pending_payout_amount']:
                cell.alignment = right_align
                cell.number_format = '$#,##0.00'
                
                # Bold key metrics
                if is_summary_metric or is_check_metric:
                    cell.font = Font(bold=True)
            else:
                cell.alignment = center_align
    
    # Auto-fit all columns using Excel's auto-sizing
    for column in worksheet.columns:
        column_letter = None
        max_length = 0
        
        for cell in column:
            # Skip merged cells
            if hasattr(cell, 'column_letter'):
                column_letter = cell.column_letter
            elif hasattr(cell, 'column'):
                column_letter = get_column_letter(cell.column)
            else:
                continue
                
            try:
                if cell.value is not None:
                    # Calculate length considering currency formatting
                    if isinstance(cell.value, (int, float)) and cell.number_format.startswith('$'):
                        # Account for currency formatting
                        cell_length = len(f"${cell.value:,.2f}")
                    else:
                        cell_length = len(str(cell.value))
                    
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        if column_letter:
            # Add padding and set reasonable min/max widths
            adjusted_width = min(max(max_length + 3, 8), 20)
            worksheet.column_dimensions[column_letter].width = adjusted_width

def format_excel_worksheet_xlsxwriter(workbook, worksheet, df):
    """Format Excel worksheet using xlsxwriter (fallback)"""
    
    # Define formats
    header_format = workbook.add_format({
        'bold': True,
        'font_color': 'white',
        'bg_color': '#2E4057',
        'align': 'center',
        'valign': 'vcenter',
        'border': 1
    })
    
    high_diff_format = workbook.add_format({
        'bg_color': '#FFCDD2',
        'border': 1,
        'num_format': '$#,##0.00'
    })
    
    # Add basic formatting (xlsxwriter is more complex to implement fully)
    worksheet.write(0, 0, "🔍 Payment Reconciliation Mismatches Analysis (Shop Timezone)", 
                   workbook.add_format({'bold': True, 'font_size': 16}))
    worksheet.write(1, 0, "🕐 Data aligned to shop timezone for consistency with Shopify reports", 
                   workbook.add_format({'italic': True, 'font_size': 10}))

def open_excel_file(filename):
    """Open Excel file in default application"""
    
    try:
        abs_path = os.path.abspath(filename)
        
        if os.name == 'nt':  # Windows
            # Try different methods for Windows
            try:
                # Method 1: Use os.startfile (original method)
                os.startfile(abs_path)
                print(f"✅ Opened {filename} in Excel")
                return
            except Exception as e1:
                print(f"⚠️  os.startfile failed: {e1}")
                
                # Method 2: Use subprocess with PowerShell
                try:
                    subprocess.run(['powershell', 'Start-Process', f'"{abs_path}"'], check=True)
                    print(f"✅ Opened {filename} in Excel (via PowerShell)")
                    return
                except Exception as e2:
                    print(f"⚠️  PowerShell Start-Process failed: {e2}")
                    
                    # Method 3: Use subprocess with cmd
                    try:
                        subprocess.run(['cmd', '/c', 'start', '', f'"{abs_path}"'], check=True)
                        print(f"✅ Opened {filename} in Excel (via cmd)")
                        return
                    except Exception as e3:
                        print(f"⚠️  cmd start failed: {e3}")
                        
        elif os.name == 'posix':  # macOS and Linux
            subprocess.call(['open', abs_path])  # macOS
            # subprocess.call(['xdg-open', abs_path])  # Linux
            print(f"✅ Opened {filename} in Excel")
            return
        
        # If all methods fail, provide manual instructions
        print(f"⚠️  Could not auto-open Excel file")
        print(f"📂 Please manually open: {abs_path}")
        
    except Exception as e:
        print(f"❌ Error opening Excel file: {e}")
        print(f"📂 Please manually open: {os.path.abspath(filename)}")

def create_utc_excel_table():
    """Create UTC timezone version of Excel analysis"""
    
    try:
        # Load transposed reconciliation data (UTC timezone for payout reconciliation)
        df = pd.read_csv('transposed_reconciliation_utc.csv', index_col=0)
        print("📊 BEAUTIFUL EXCEL MISMATCH ANALYSIS (UTC TIMEZONE)")
        print("=" * 60)
        print("🕐 Using UTC timezone data for payout reconciliation accuracy")
        
        # Get mismatch dates
        mismatch_row = df.loc['mismatch'] if 'mismatch' in df.index else None
        if mismatch_row is None:
            print("❌ No 'mismatch' row found in UTC data")
            return
            
        mismatch_dates = [col for col in df.columns if str(mismatch_row[col]).lower() == 'true']
        
        if not mismatch_dates:
            print("✅ No mismatches found in UTC data! All dates reconcile properly.")
            return
        
        # Apply date filtering and sorting
        print(f"📊 Found {len(mismatch_dates)} mismatched dates in UTC data")
        filtered_dates = filter_and_sort_dates(mismatch_dates)
        
        if not filtered_dates:
            print("❌ No dates remain after filtering!")
            return
        
        print("📝 Creating UTC timezone Excel file...")
        
        # Use filtered dates
        mismatch_dates = filtered_dates
        
        # Use the same logic as the shop timezone version
        focus_rows = [
            'order_count', 'gross_sales', 'subtotal', 'discounts', 'tax', 'shipping', 'tips',
            'net_payment', 'refunds', 'shopify_payments', 'gift_card', 'cash', 'manual',
            'payout_type_adjustment', 'payout_type_chargeback', 'payout_type_refund',
            'payout_amount', 'reconciliation_difference', 'pending_payout_amount', 'pending_payout_count'
        ]
        
        # Process the data similar to shop timezone version...
        available_rows = [row for row in focus_rows if row in df.index]
        mismatch_df = df.loc[available_rows, mismatch_dates]
        
        # Convert to numeric and handle calculations
        for date_col in mismatch_dates:
            if date_col in mismatch_df.columns:
                for idx in mismatch_df.index:
                    try:
                        mismatch_df.loc[idx, date_col] = pd.to_numeric(mismatch_df.loc[idx, date_col], errors='coerce')
                        if pd.isna(mismatch_df.loc[idx, date_col]):
                            mismatch_df.loc[idx, date_col] = 0
                    except:
                        mismatch_df.loc[idx, date_col] = 0
        
        # Add calculated rows
        sales_total_row = (mismatch_df.loc['gross_sales'] - mismatch_df.loc['discounts'] + 
                          mismatch_df.loc['tax'] + mismatch_df.loc['shipping'] + mismatch_df.loc['tips'])
        expected_from_sales_row = mismatch_df.loc['net_payment'] + mismatch_df.loc['refunds']
        sales_balance_check_row = sales_total_row - expected_from_sales_row
        
        gateway_payments_row = (mismatch_df.loc['shopify_payments'] + mismatch_df.loc['gift_card'] + 
                               mismatch_df.loc['cash'] + mismatch_df.loc['manual'])
        payment_adjustments_row = (mismatch_df.loc['payout_type_adjustment'] + 
                                  mismatch_df.loc['payout_type_chargeback'] + 
                                  mismatch_df.loc['payout_type_refund'])
        
        # For proper reconciliation: compare total shopify receipts with expected payout before fees
        # Total Shopify receipts = shopify_payments + payout_refunds (since refunds were originally part of payments)
        payout_charges_row = mismatch_df.loc['payout_amount']  # Charges from payout data (before fees)
        payout_refunds_row = abs(mismatch_df.loc['payout_type_refund'])  # Refunds (make positive for clarity)
        expected_payout_before_fees_row = payout_charges_row  # This is the expected payout before fees
        
        # Total Shopify payment receipts includes the original payments plus refunds
        total_shopify_receipts_row = mismatch_df.loc['shopify_payments'] + payout_refunds_row
        
        total_gateway_receipts_row = gateway_payments_row + payment_adjustments_row
        payment_difference_row = total_shopify_receipts_row - expected_payout_before_fees_row
        
        # Add rows to DataFrame
        mismatch_df.loc['sales_total'] = sales_total_row
        mismatch_df.loc['expected_from_sales'] = expected_from_sales_row  
        mismatch_df.loc['sales_balance_check'] = sales_balance_check_row
        mismatch_df.loc['empty_row'] = 0
        mismatch_df.loc['gateway_payments'] = gateway_payments_row
        mismatch_df.loc['payment_adjustments'] = payment_adjustments_row
        mismatch_df.loc['total_gateway_receipts'] = total_gateway_receipts_row
        mismatch_df.loc['payout_charges'] = payout_charges_row
        mismatch_df.loc['payout_refunds_amount'] = payout_refunds_row
        mismatch_df.loc['total_shopify_receipts'] = total_shopify_receipts_row
        mismatch_df.loc['expected_payout_before_fees'] = expected_payout_before_fees_row
        mismatch_df.loc['payment_difference'] = payment_difference_row
        
        # Reorder for Excel layout
        excel_rows = [
            'order_count', 'gross_sales', 'subtotal', 'discounts', 'tax', 'shipping', 'tips',
            'sales_total', 'net_payment', 'refunds', 'expected_from_sales', 'sales_balance_check',
            'empty_row', 'shopify_payments', 'gift_card', 'cash', 'manual', 'gateway_payments',
            'payout_type_adjustment', 'payout_type_chargeback', 'payout_type_refund', 'payment_adjustments',
            'total_gateway_receipts', 'empty_row', 'payout_amount', 'payout_charges', 'payout_refunds_amount',
            'total_shopify_receipts', 'expected_payout_before_fees', 'payment_difference', 'reconciliation_difference',
            'pending_payout_amount', 'pending_payout_count'
        ]
        
        available_rows = [row for row in excel_rows if row in mismatch_df.index]
        mismatch_dates_sorted = mismatch_dates  # Already sorted by filter function
        excel_df = mismatch_df.loc[available_rows, mismatch_dates_sorted].copy()
        
        # Create Excel file with UTC timezone formatting
        from datetime import datetime
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f'Payment_Reconciliation_Mismatches_UTC_{timestamp}.xlsx'
        
        try:
            create_formatted_excel_utc(excel_df, 'Payment_Reconciliation_Mismatches_UTC.xlsx')
            excel_filename = 'Payment_Reconciliation_Mismatches_UTC.xlsx'
        except PermissionError:
            print("⚠️  Previous UTC Excel file is open, creating new timestamped version...")
            create_formatted_excel_utc(excel_df, excel_filename)
        
        print(f"✅ Created UTC version: {excel_filename}")
        print("🚀 Opening UTC Excel file...")
        
        # Open in Excel
        open_excel_file(excel_filename)
        
    except FileNotFoundError:
        print("❌ transposed_reconciliation_utc.csv not found")
        print("💡 Run reconcile_payouts.py first to generate the UTC timezone data")
        print("🕐 This analysis uses UTC timezone data for payout reconciliation")
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()

def create_formatted_excel_utc(df, filename):
    """Create beautifully formatted UTC timezone Excel file"""
    
    try:
        # Try to use openpyxl for better formatting
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            # Write main data
            df.to_excel(writer, sheet_name='UTC Mismatch Analysis', startrow=2)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['UTC Mismatch Analysis']
            
            # Apply UTC-specific formatting
            format_excel_worksheet_utc(workbook, worksheet, df)
            
    except ImportError:
        # Fallback to xlsxwriter
        try:
            with pd.ExcelWriter(filename, engine='xlsxwriter') as writer:
                df.to_excel(writer, sheet_name='UTC Mismatch Analysis', startrow=2)
                
                workbook = writer.book
                worksheet = writer.sheets['UTC Mismatch Analysis']
                
                format_excel_worksheet_xlsxwriter_utc(workbook, worksheet, df)
                
        except ImportError:
            # Basic Excel export
            df.to_excel(filename, sheet_name='UTC Mismatch Analysis')
            print("⚠️  Basic UTC Excel file created (install openpyxl or xlsxwriter for formatting)")

def format_excel_worksheet_utc(workbook, worksheet, df):
    """Format Excel worksheet for UTC timezone analysis"""
    
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    # Define colors (same as main function)
    header_fill = PatternFill(start_color="2E4057", end_color="2E4057", fill_type="solid")
    high_diff_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    medium_diff_fill = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
    low_diff_fill = PatternFill(start_color="F3E5F5", end_color="F3E5F5", fill_type="solid")
    
    # Define fonts
    header_font = Font(color="FFFFFF", bold=True, size=11)
    bold_font = Font(bold=True)
    
    # Define alignment
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    
    # Define border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'), 
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Add title with UTC indication
    title_cell = worksheet['A1']
    title_cell.value = "🔍 Payment Reconciliation Mismatches Analysis (UTC Timezone)"
    title_cell.font = Font(size=16, bold=True, color="2E4057")
    
    # Add timezone info in row 2
    timezone_cell = worksheet['A2']
    timezone_cell.value = "🕐 Data aligned to UTC timezone for accurate payout reconciliation"
    timezone_cell.font = Font(size=10, italic=True, color="666666")
    
    # Merge title across all columns
    last_col = get_column_letter(len(df.columns) + 1)
    worksheet.merge_cells(f'A1:{last_col}1')
    worksheet.merge_cells(f'A2:{last_col}2')
    
    # Format date headers (row 3, columns B onwards)
    header_row = 3
    for col_num, date_col in enumerate(df.columns, 2):  # Start at column B
        cell = worksheet.cell(row=header_row, column=col_num)
        cell.value = date_col
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border
    
    # Format metric row headers (column A, starting row 4)
    metric_labels = {
        # Sales Analysis
        'order_count': 'Order Count',
        'gross_sales': 'Gross Sales',
        'subtotal': 'Subtotal',
        'discounts': 'Discounts',
        'tax': 'Tax',
        'shipping': 'Shipping',
        'tips': 'Tips',
        'sales_total': '📊 Sales Total (Gross - Disc + Tax + Ship + Tips)',
        'net_payment': 'Net Payment Expected',
        'refunds': 'Refunds',
        'expected_from_sales': '🔍 Expected from Sales (Net + Refunds)',
        'sales_balance_check': '⚠️ Sales Balance Check',
        'empty_row': '',
        # Payment Gateways
        'shopify_payments': 'Shopify Payments',
        'gift_card': 'Gift Cards',
        'cash': 'Cash',
        'manual': 'Manual',
        'gateway_payments': '💳 Gateway Payments Subtotal',
        'payout_type_adjustment': 'Payout Adjustments',
        'payout_type_chargeback': 'Chargebacks',
        'payout_type_refund': 'Payout Refunds',
        'payment_adjustments': '🔧 Payment Adjustments Subtotal',
        'total_gateway_receipts': '📈 Total Gateway Receipts',
        # Payout Reconciliation
        'payout_amount': 'Payout Charges',
        'payout_charges': 'Payout Charges (same as above)',
        'payout_refunds_amount': 'Payout Refunds (positive)',
        'total_shopify_receipts': '💰 Total Shopify Payment Receipts (Payments + Refunds)',
        'expected_payout_before_fees': '🧮 Expected Payout Before Fees',
        'payment_difference': '❗ Payment Difference (Receipts - Expected)',
        'reconciliation_difference': 'Reconciliation Difference',
        'pending_payout_amount': 'Pending Amount',
        'pending_payout_count': 'Pending Count'
    }
    
    # Format metric name column header
    metric_header = worksheet.cell(row=header_row, column=1)
    metric_header.value = 'Metric'
    metric_header.fill = header_fill
    metric_header.font = header_font
    metric_header.alignment = right_align  # Changed from center_align to right_align
    metric_header.border = thin_border
    
    # Format data rows
    for row_num, (metric, row_data) in enumerate(df.iterrows(), header_row + 1):
        # Handle empty spacer row
        if metric == 'empty_row':
            for col_num in range(1, len(df.columns) + 2):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = ''
                cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            continue
            
        # Metric name column
        metric_cell = worksheet.cell(row=row_num, column=1)
        metric_cell.value = metric_labels.get(metric, metric.replace('_', ' ').title())
        metric_cell.alignment = right_align  # Changed from center_align to right_align
        metric_cell.border = thin_border
        
        # Determine special formatting for different row types
        is_summary_metric = metric in ['sales_total', 'expected_from_sales', 'gateway_payments', 
                                     'payment_adjustments', 'total_gateway_receipts', 'expected_payout_net']
        is_check_metric = metric in ['sales_balance_check', 'payment_difference']
        is_section_header = metric in ['order_count']  # First row of each section
        
        # Set metric cell formatting
        if is_summary_metric:
            metric_cell.font = Font(bold=True, color="1F4E79")
            metric_cell.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
        elif is_check_metric:
            metric_cell.font = Font(bold=True, color="C5504B")  
            metric_cell.fill = PatternFill(start_color="FFEEEE", end_color="FFEEEE", fill_type="solid")
        else:
            metric_cell.font = bold_font
            metric_cell.fill = PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        # Format data columns
        for col_num, date_col in enumerate(df.columns, 2):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell_value = row_data[date_col]
            cell.value = cell_value
            cell.border = thin_border
            
            # Determine cell color based on different criteria
            if metric == 'payment_difference':
                payment_diff = abs(float(cell_value)) if cell_value != 0 else 0
                if payment_diff > 100:
                    cell.fill = high_diff_fill
                elif payment_diff > 50:
                    cell.fill = medium_diff_fill
                else:
                    cell.fill = low_diff_fill
            elif metric == 'sales_balance_check':
                # Highlight sales balance issues
                balance_diff = abs(float(cell_value)) if cell_value != 0 else 0
                if balance_diff > 1:  # More than $1 difference
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            elif is_summary_metric:
                cell.fill = PatternFill(start_color="E7F3FF", end_color="E7F3FF", fill_type="solid")
            elif is_check_metric:
                cell.fill = PatternFill(start_color="FFEEEE", end_color="FFEEEE", fill_type="solid")
            
            # Number formatting and alignment
            if metric in ['order_count', 'pending_payout_count']:
                cell.alignment = center_align
                cell.number_format = '0'
            elif metric in ['gross_sales', 'subtotal', 'discounts', 'tax', 'shipping', 'tips', 'sales_total',
                          'net_payment', 'refunds', 'expected_from_sales', 'sales_balance_check',
                          'shopify_payments', 'gift_card', 'cash', 'manual', 'gateway_payments',
                          'payout_type_adjustment', 'payout_type_chargeback', 'payout_type_refund', 'payment_adjustments',
                          'total_gateway_receipts', 'payment_difference',
                          'payout_amount', 'payout_charges', 'payout_refunds_amount', 'expected_payout_net',
                          'total_shopify_receipts', 'expected_payout_before_fees',  # Added new metrics
                          'reconciliation_difference', 'pending_payout_amount']:
                cell.alignment = right_align
                cell.number_format = '$#,##0.00'
                
                # Bold key metrics
                if is_summary_metric or is_check_metric:
                    cell.font = Font(bold=True)
            else:
                cell.alignment = center_align
    
    # Auto-fit all columns using Excel's auto-sizing
    for column in worksheet.columns:
        column_letter = None
        max_length = 0
        
        for cell in column:
            # Skip merged cells
            if hasattr(cell, 'column_letter'):
                column_letter = cell.column_letter
            elif hasattr(cell, 'column'):
                column_letter = get_column_letter(cell.column)
            else:
                continue
                
            try:
                if cell.value is not None:
                    # Calculate length considering currency formatting
                    if isinstance(cell.value, (int, float)) and cell.number_format.startswith('$'):
                        # Account for currency formatting
                        cell_length = len(f"${cell.value:,.2f}")
                    else:
                        cell_length = len(str(cell.value))
                    
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        if column_letter:
            # Add padding and set reasonable min/max widths
            adjusted_width = min(max(max_length + 3, 8), 20)
            worksheet.column_dimensions[column_letter].width = adjusted_width

def format_excel_worksheet_xlsxwriter_utc(workbook, worksheet, df):
    """Format Excel worksheet using xlsxwriter for UTC timezone (fallback)"""
    
    # Add basic formatting with UTC indication
    worksheet.write(0, 0, "🔍 Payment Reconciliation Mismatches Analysis (UTC Timezone)", 
                   workbook.add_format({'bold': True, 'font_size': 16}))
    worksheet.write(1, 0, "🕐 Data aligned to UTC timezone for accurate payout reconciliation", 
                   workbook.add_format({'italic': True, 'font_size': 10}))

def main():
    """Main function with interactive menu"""
    
    while True:
        print("\n🔍 BEAUTIFUL EXCEL MISMATCH VIEWER")
        print("=" * 40)
        print("1. Generate Shop Timezone Excel Analysis")
        print("2. Generate UTC Timezone Excel Analysis")
        print("3. Configure Date Settings")
        print("4. Show Current Configuration")
        print("5. Test Excel File Opening")
        print("6. Exit")
        
        choice = input("\nEnter your choice (1-6): ").strip()
        
        if choice == "1":
            print("\n🏪 Creating Shop Timezone Excel Analysis...")
            create_beautiful_excel_table()
        elif choice == "2":
            print("\n🌍 Creating UTC Timezone Excel Analysis...")
            create_utc_excel_table()
        elif choice == "3":
            configure_date_settings()
        elif choice == "4":
            show_date_configuration()
        elif choice == "5":
            # Test opening the most recent Excel file
            import glob
            excel_files = glob.glob("Payment_Reconciliation_Mismatches*.xlsx")
            if excel_files:
                # Sort by modification time, newest first
                excel_files.sort(key=os.path.getmtime, reverse=True)
                test_file = excel_files[0]
                print(f"\n🧪 Testing Excel file opening with: {test_file}")
                open_excel_file(test_file)
            else:
                print("\n❌ No Excel files found to test")
                print("💡 Run option 1 or 2 first to create Excel files")
        elif choice == "6":
            print("\n👋 Goodbye!")
            break
        else:
            print("\n❌ Invalid choice. Please try again.")

if __name__ == "__main__":
    main()
